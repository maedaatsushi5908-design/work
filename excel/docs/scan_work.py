#!/usr/bin/env python3
"""工事フォルダを調べて、M_Hasai の COL_MAP の下書きを出す

新しい工事の数量計算書一式を excel/works/<名前>/ に入れたら、これを流す。
総括表の見出し・シート名・舗装版破砕ブロックの有無を突き合わせ、
マクロの先頭に貼る COL_MAP を作る。

    python3 excel/docs/scan_work.py excel/works/02
"""
import os
import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter as gl

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import verify_hasai as H            # noqa: E402  判定はマクロと同じものを使う


def header_of(ws, col):
    for r in range(1, 13):
        v = ws.cell(r, col).value
        if v not in (None, "") and H.norm(v) and not str(v).startswith("="):
            return " ".join(str(v).split())
    return ""


def kei_of(ws, col):
    """系統＝見出しの括弧より前"""
    h = header_of(ws, col)
    if not h:
        return ""
    for br in ("（", "("):
        p = h.find(br)
        if p > 0:
            h = h[:p]
            break
    return h.replace(" ", "").replace("　", "").strip()


def lead(sn):
    """シート名の括弧より前  仮配（舗 → 仮配"""
    t = H.norm(sn)
    for ch in ("(", "（"):
        p = t.find(ch)
        if p > 0:
            return t[:p]
    return t


def find_target(wb):
    """いちばん左の工種名が「舗装版破砕」の行があるシートを探す"""
    hits = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, min(ws.max_row, 250) + 1):
            for c in range(1, 10):
                v = H.merged_value(ws, r, c)
                if v in (None, "") or str(v).startswith("="):
                    continue
                if str(v).strip():
                    if H.norm(H.SECTION_LABEL) in H.norm(v):
                        hits.append((sn, r))
                    break            # その行のいちばん左だけ見る
            if hits and hits[-1][0] == sn:
                break
    return hits


def load_all(folder):
    """フォルダ内の xlsx/xlsm を全部開く"""
    out = []
    for f in sorted(os.listdir(folder)):
        if not f.lower().endswith((".xlsx", ".xlsm")) or f.startswith("~"):
            continue
        try:
            out.append((f, openpyxl.load_workbook(os.path.join(folder, f))))
        except Exception as e:                                  # noqa: BLE001
            print(f"--- {f}  読めません: {e}")
    return out


def assign(cands, order, sheets_of_kei, kei_of_col):
    """列 → 転記元シートを決める

    同じ系統の列数と転記元シート数が同じなら、並び順どおりに組む。
    様式では 1枠目=新設 / 3枠目=取替・撤去 のように枠の意味が決まっていて、
    シート名の数字（管工（舗50 の 50）は口径ではなく枠の名前だから。
    数が合わないときは、候補の少ない列から順に決めて使った分を外す。
    """
    chosen, how = {}, {}
    done = set()
    for kei, sheets in sheets_of_kei.items():
        cls = [c for c in order if kei_of_col.get(c) == kei and cands.get(c)]
        if len(cls) >= 2 and len(cls) == len(sheets):
            for cl, sn in zip(cls, sheets):
                chosen[cl] = sn
                how[cl] = "並び順"
                done.add(cl)

    used = set(chosen.values())
    todo = {c: list(v) for c, v in cands.items() if v and c not in done}
    while todo:
        best = min(todo, key=lambda c: (len([x for x in todo[c] if x not in used]),
                                        order.index(c)))
        left = [x for x in todo[best] if x not in used]
        del todo[best]
        if not left:
            continue
        chosen[best] = left[0]
        how[best] = "見出しの口径" if len(left) == 1 else "★候補が複数"
        used.add(left[0])
    return chosen, how


def scan(folder):
    xls = sorted(f for f in os.listdir(folder) if f.lower().endswith(".xls"))
    if xls:
        print(f"※ .xls（97-2003 形式）は飛ばします: {', '.join(xls)}\n")

    books = load_all(folder)

    # --- 転記元（舗装版破砕のブロックを持つシート）をブック横断で集める ---
    srcs = []                      # (ファイル名, シート名, ブロック)
    for fn, wb in books:
        H._cache.clear()
        for sn in wb.sheetnames:
            b = H.find_block(wb, sn)
            if b:
                srcs.append((fn, sn, b))
    if not srcs:
        print("舗装版破砕のブロックを持つシートが見つかりませんでした。")
        return

    print("=" * 74)
    print(f"■ 転記元シート {len(srcs)} 枚")
    cur = None
    for fn, sn, b in srcs:
        if fn != cur:
            print(f"  {fn}")
            cur = fn
        r0, r1, at, a_s, ct, cs = b
        co = f"Co={gl(ct)}/{gl(cs)}" if cs else "Co なし"
        print(f"    {sn:<16} {r0:>3}-{r1:<3}行  As={gl(at)}/{gl(a_s):<3} {co}")

    block_names = {sn for _f, sn, _b in srcs}
    src_files = sorted({fn for fn, _s, _b in srcs})
    # 全ブックのシート名。「給水2度」のように総括表とは別のブックに
    # 同名シートがあることがあるので、横断で見る
    all_sheets = {sn for _f, w in books for sn in w.sheetnames}

    # --- 総括表を探して突き合わせる ---
    for fn, wb in books:
        for tsn, trow in find_target(wb):
            if tsn in block_names:
                continue                       # 転記元シート自身
            ws = wb[tsn]
            fc = 0
            for c in range(1, 40):
                if kei_of(ws, c):
                    fc = c
                    break
            if not fc:
                continue

            cands, info, order, kei_of_col = {}, {}, [], {}
            for c in range(fc, min(ws.max_column, 60) + 1):
                k = kei_of(ws, c)
                if not k:
                    continue
                yellow = sum(1 for r in range(trow, min(trow + 90, ws.max_row + 1))
                             if H.is_input_cell(ws, r, c))
                if yellow == 0:
                    continue
                cl = gl(c)
                order.append(cl)
                kei_of_col[cl] = k
                info[cl] = (header_of(ws, c), k, yellow)

                exact = [sn for _f, sn, _b in srcs if H.norm(sn) == H.norm(k)]
                if exact:
                    cands[cl] = [exact[0]]
                    continue
                if any(H.norm(sn) == H.norm(k) for sn in all_sheets):
                    cands[cl] = []              # 同名だがブロックが無い＝対象外
                    continue
                nums = re.findall(r"\d+", header_of(ws, c))
                lst = []
                for _f, sn, _b in srcs:
                    if H.norm(k).startswith(lead(sn)):
                        tail = re.search(r"(\d+)$", sn)
                        if tail is None or tail.group(1) in nums:
                            lst.append(sn)
                cands[cl] = lst

            if len([c for c in cands if cands[c]]) < 2:
                continue

            # 系統ごとの転記元シート（ブックの並び順のまま）
            sheets_of_kei = {}
            for cl in order:
                k = kei_of_col[cl]
                if k in sheets_of_kei:
                    continue
                lst = [sn for _f, sn, _b in srcs if H.norm(k).startswith(lead(sn))
                       and H.norm(sn) != H.norm(k)]
                if lst:
                    sheets_of_kei[k] = lst
            chosen, how = assign(cands, order, sheets_of_kei, kei_of_col)

            print("\n" + "=" * 74)
            print(f"▼ 総括表「{tsn}」　{fn}　{H.SECTION_LABEL} は {trow} 行目から")
            print(f"  {'列':<3} {'見出し':<32} {'系統':<9} {'黄':>3}  転記元         決め方")
            print("  " + "-" * 84)
            proposal, unsure = [], []
            for cl in order:
                hdr, k, yellow = info[cl]
                if cl in chosen:
                    sn = chosen[cl]
                    proposal.append((cl, sn))
                    note = how[cl]
                    if note.startswith("★"):
                        unsure.append(cl)
                elif cands.get(cl) == []:
                    sn, note = "—", "同名シートに破砕のブロックが無い → 対象外"
                else:
                    sn, note = "—", "★候補なし"
                    unsure.append(cl)
                print(f"  {cl:<3} {hdr[:32]:<32} {k:<9} {yellow:>3}  {sn:<14} {note}")

            if proposal:
                print("\n  --- マクロ先頭に貼る下書き ---")
                print(f'  Private Const TARGET_SHEET As String = "{tsn}"')
                other = [f for f in src_files if f != fn]
                if other:
                    print(f'  \' 転記元は別のブック: {", ".join(other)}')
                print('  Private Const COL_MAP As String = _')
                for i, (cl, sn) in enumerate(proposal):
                    last = (i == len(proposal) - 1)
                    print(f'      "{cl}={sn}{"" if last else "|"}"{"" if last else " & _"}')
                if unsure:
                    print(f"\n  ★ {', '.join(unsure)} 列は決め切れていません。"
                          "見出しとシートを見比べて確かめてください。")
    print()


def main(argv):
    if not argv:
        print(__doc__)
        return 1
    folder = argv[0]
    if not os.path.isdir(folder):
        print("フォルダがありません:", folder)
        return 1
    if not [f for f in os.listdir(folder) if f.lower().endswith((".xls", ".xlsx", ".xlsm"))]:
        print(f"{folder} に Excel ファイルがありません。")
        return 1
    scan(folder)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
