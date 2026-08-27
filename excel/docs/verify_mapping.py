#!/usr/bin/env python3
"""転記定義の検証スクリプト

VBA の M_Engine と同じ解決ロジックを移植したもの。
excel/works/01_higashishirakawa/ の実ファイルに対して、定義どおりの行・列に到達し、
総括表の現在値と一致するかを確認する。

    cd excel/works/01_higashishirakawa && python3 ../../docs/verify_mapping.py

VBA を書き換えたら、こちらのロジックも合わせて更新すること。
"""
import re
import unicodedata

import openpyxl
import xlrd


def norm(s):
    """全角→半角・空白除去・φ除去・大文字化。VBA の Norm と同じ"""
    if s is None:
        return ""
    t = unicodedata.normalize("NFKC", str(s))
    for ch in " 　\t\r\nφΦfF":
        t = t.replace(ch, "")
    return t.upper()


def extract_num(s):
    """最初の数字の並びを数値で返す。VBA の ExtractNum と同じ"""
    if s is None:
        return -1
    m = re.search(r"\d+", unicodedata.normalize("NFKC", str(s)))
    return float(m.group()) if m else -1


def key_matches(target, key, mode):
    if not key:
        return False
    if mode == "口径":
        a, b = extract_num(target), extract_num(key)
        return a >= 0 and b >= 0 and a == b
    if mode == "完全":
        return norm(target) == norm(key)
    nt, nk = norm(target), norm(key)
    return len(nt) > 0 and nk in nt


class Sheet:
    """xlsx / xls を同じ形で扱う"""

    def __init__(self, path, name):
        self.grid = {}
        if path.endswith("x"):
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = next(w for w in wb.worksheets if norm(w.title) == norm(name))
            for row in ws.iter_rows():
                for c in row:
                    if c.value is not None:
                        self.grid[(c.row, c.column)] = c.value
            self.nrows, self.ncols = ws.max_row, ws.max_column
        else:
            b = xlrd.open_workbook(path)
            sh = next(b.sheet_by_name(n) for n in b.sheet_names() if norm(n) == norm(name))
            for r in range(sh.nrows):
                for c in range(sh.ncols):
                    v = sh.cell_value(r, c)
                    if v not in ("", None):
                        self.grid[(r + 1, c + 1)] = v
            self.nrows, self.ncols = sh.nrows, sh.ncols

    def cell(self, r, c):
        return self.grid.get((r, c))


def col_num(letter):
    if not letter:
        return 0
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def find_row(sh, k1c, k1, k1m, k2c, k2, k2m, off):
    """キー列は上方向に値を補完してから照合する。VBA の FindTargetRow と同じ"""
    k1c, k2c = col_num(k1c), col_num(k2c)
    cur1 = ""
    for r in range(1, sh.nrows + 1):
        if k1c:
            v = sh.cell(r, k1c)
            if norm(v):
                cur1 = str(v)
        ok1 = True if (not k1c or not k1) else key_matches(cur1, k1, k1m)
        if not ok1:
            continue
        ok2 = True if (not k2c or not k2) else key_matches(sh.cell(r, k2c), k2, k2m)
        if ok2:
            return r + off
    return 0


def find_col(sh, valcol, hdrrow, hdrkey):
    if valcol:
        return col_num(valcol)
    if not hdrkey:
        return 0
    mode = "口径" if str(hdrkey).isdigit() else "部分"
    if hdrrow:
        for c in range(1, sh.ncols + 1):
            if key_matches(sh.cell(hdrrow, c), hdrkey, mode):
                return c
    for r in range(1, min(40, sh.nrows) + 1):
        for c in range(1, sh.ncols + 1):
            if key_matches(sh.cell(r, c), hdrkey, mode):
                return c
    return 0


# 名称, 摘要, 転記先, ファイル, シート, キー1列, キー1, 方式,
# キー2列, キー2, 方式, 行オフセット, 値列, 見出し行, 見出しキー, 丸め, 期待値
CHUTETSU = "02_chutetsukan.xls"
KONKYO = "05-1_dokou_konkyo.xlsx"
ENCHO = "延長集計表  採用(伏越削除)"
DEFS = [
    ("鋳鉄管据付", "200", "D10", CHUTETSU, "　印刷　", "B", "据付延長 設計書 入力数値", "部分", "", "", "", 0, "", 5, "200", 1, 41.2),
    ("鋳鉄管据付", "300", "D11", CHUTETSU, "　印刷　", "B", "据付延長 設計書 入力数値", "部分", "", "", "", 0, "", 5, "300", 1, 1.5),
    ("鋳鉄管据付", "400", "D12", CHUTETSU, "　印刷　", "B", "据付延長 設計書 入力数値", "部分", "", "", "", 0, "", 5, "400", 1, 638.1),
    ("GX継手（直部）", "200", "J32", CHUTETSU, "　印刷　", "C", "GX形 直 管", "部分", "D", "200", "口径", 0, "F", 0, "", 0, 8.0),
    ("GX継手（直部）", "400", "J34", CHUTETSU, "　印刷　", "C", "GX形 直 管", "部分", "D", "400", "口径", 0, "F", 0, "", 0, 114.0),
    ("GX継手（異形部）", "200", "J38", CHUTETSU, "　印刷　", "C", "GX形 異形管接合材", "部分", "D", "200", "口径", 0, "F", 0, "", 0, 18.0),
    ("GX継手（異形部）", "300", "J39", CHUTETSU, "　印刷　", "C", "GX形 異形管接合材", "部分", "D", "300", "口径", 0, "F", 0, "", 0, 3.0),
    ("GX継手（異形部）", "400", "J40", CHUTETSU, "　印刷　", "C", "GX形 異形管接合材", "部分", "D", "400", "口径", 0, "F", 0, "", 0, 100.0),
    ("GX継手（特殊押輪部）", "400", "J45", CHUTETSU, "　印刷　", "C", "GX形 特殊押輪", "部分", "D", "400", "口径", 0, "F", 0, "", 0, 2.0),
    ("挿口加工費（GX形）", "200", "D139", CHUTETSU, "　印刷　", "C", "GX形 挿口リング", "部分", "D", "200", "口径", 0, "F", 0, "", 0, 7.0),
    ("挿口加工費（GX形）", "400", "D140", CHUTETSU, "　印刷　", "C", "GX形 挿口リング", "部分", "D", "400", "口径", 0, "F", 0, "", 0, 50.0),
    ("鋼管据付", "80", "D18", KONKYO, ENCHO, "G", "SP80A", "完全", "", "", "", 0, "K", 0, "", 1, 1.1),
    ("鋼管据付", "300", "D22", KONKYO, ENCHO, "G", "SP300A", "完全", "", "", "", 0, "K", 0, "", 1, 1.0),
    ("鋼管据付", "400", "D23", KONKYO, ENCHO, "G", "SP400A", "完全", "", "", "", 0, "K", 0, "", 1, 459.3),
    ("鋼管撤去", "250", "D44", KONKYO, ENCHO, "G", "SP250A", "完全", "", "", "", 0, "L", 0, "", 1, 2.5),
    ("鋼管撤去", "300", "D45", KONKYO, ENCHO, "G", "SP300A", "完全", "", "", "", 0, "L", 0, "", 1, 1023.4),
    ("鋼管撤去", "400", "D46", KONKYO, ENCHO, "G", "SP400A", "完全", "", "", "", 0, "L", 0, "", 1, 459.2),
    ("電送管据付", "80（FEP)", "J62", KONKYO, ENCHO, "O", "布設延長合計", "部分", "", "", "", 0, "P", 0, "", 1, 2158.8),
    ("電送管撤去", "82（VE)", "J64", KONKYO, ENCHO, "O", "撤去延長合計", "部分", "", "", "", 0, "P", 0, "", 1, 2840.4),
    ("鋳鉄管切断 （新　管）", "200", "J86", "03_kirikan.xlsx", "GX　200", "B", "合 計", "部分", "", "", "", 1, "U", 0, "", 0, 7.0),
    ("鋳鉄管切断 （新　管）", "400", "J88", "03_kirikan.xlsx", "GX　400 (4)", "B", "合 計", "部分", "", "", "", 1, "U", 0, "", 0, 48.0),
    ("ステンレス鋼管 現場溶接", "80", "D60", "08_koukan.xls", "工事数量表", "A", "電気溶接", "部分", "L", "80A", "部分", 0, "AO", 0, "", 0, 1.0),
    ("ステンレス鋼管 現場溶接", "300", "D61", "08_koukan.xls", "工事数量表", "A", "電気溶接", "部分", "L", "300A", "部分", 0, "AO", 0, "", 0, 3.0),
    ("ステンレス鋼管 現場溶接", "400", "D62", "08_koukan.xls", "工事数量表", "A", "電気溶接", "部分", "L", "400A", "部分", 0, "AO", 0, "", 0, 149.0),
    ("鋼管現場溶接", "300", "D64", "08_koukan.xls", "工事数量表", "A", "閉塞蓋設置", "部分", "L", "300", "口径", 0, "AO", 0, "", 0, 2.0),
    ("鋼管現場溶接", "400", "D65", "08_koukan.xls", "工事数量表", "A", "閉塞蓋設置", "部分", "L", "400", "口径", 0, "AO", 0, "", 0, 2.0),
]


def main():
    cache = {}
    ok = ng = 0
    for (nm, sp, dest, f, sn, k1c, k1, k1m, k2c, k2, k2m, off, vc, hr, hk, dec, exp) in DEFS:
        key = (f, sn)
        if key not in cache:
            cache[key] = Sheet(f, sn)
        sh = cache[key]
        r = find_row(sh, k1c, k1, k1m, k2c, k2, k2m, off)
        c = find_col(sh, vc, hr, hk)
        v = sh.cell(r, c) if (r and c) else None
        got = round(float(v), dec) if isinstance(v, (int, float)) else None
        good = got is not None and abs(got - exp) < 1e-9
        ok, ng = (ok + 1, ng) if good else (ok, ng + 1)
        print(f"  {'OK ' if good else 'NG '}{dest:>5} {nm[:16]:16} φ{sp:8} → 行{r:>4} 列{c:>3} = {got}  (期待 {exp})")
    print(f"\n  一致 {ok} / {ok + ng}")
    return 0 if ng == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
