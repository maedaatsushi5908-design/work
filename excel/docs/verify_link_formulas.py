#!/usr/bin/env python3
"""M_Link の取り込み・展開・数式生成の検証

VBA の「リンク設定を取り込む」→「直接リンクで数式を作る」と同じ処理を
再現し、次の3点を確かめる。

  1. 既存のリンクがあるセルは、生成後も値が変わらないこと
  2. 新しく数式が入るセル（未リンクの口径列）に何が入るか
  3. 書き込み対象から外すべき行が正しく外れること

    cd excel/original && python3 ../docs/verify_link_formulas.py

VBA を書き換えたら、こちらのロジックも合わせて更新すること。
"""
import collections
import re
import sys
import unicodedata

import openpyxl
from openpyxl.utils import get_column_letter as gl

BOOK = "06_dokou_hosou.xlsx"
TARGET = "総括表（土工事）"
TERM = re.compile(r"'([^']+)'!(\$?[A-Z]{1,3}\$?[0-9]+)")


def norm(s):
    """VBA の Norm と同じ"""
    if s is None:
        return ""
    t = unicodedata.normalize("NFKC", str(s))
    for ch in " 　\t\r\nφΦfF":
        t = t.replace(ch, "")
    return t.upper()


def trail(s):
    """末尾に続く数字。VBA の TrailDigits と同じ"""
    m = re.search(r"(\d+)$", str(s))
    return m.group(1) if m else ""


def header_of(ws, col):
    """列の見出し。1〜12行目の最初の非数式セル"""
    for r in range(1, 13):
        v = ws.cell(r, col).value
        if v not in (None, "") and norm(v) and not str(v).startswith("="):
            return " ".join(str(v).split())
    return ""


def kei_of(ws, col):
    """系統＝見出しの「（」より前。VBA の KeiOf と同じ"""
    h = header_of(ws, col)
    if not h:
        return ""
    p = h.find("（")
    if p < 0:
        p = h.find("(")
    if p > 0:
        h = h[:p]
    return h.replace(" ", "").replace("　", "").strip()


def collect(ws):
    """リンクを行×系統でまとめ、展開の可否を判定する"""
    grp, keys, colinfo = {}, collections.defaultdict(set), {}
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not (isinstance(v, str) and v.startswith("=") and "!" in v):
                continue
            if "表紙" in v:
                continue
            terms = TERM.findall(v)
            if not terms or len(terms) != v.count("!"):
                continue  # 想定外の書き方は触らない

            cl, kei = gl(c.column), kei_of(ws, c.column)
            dia, sig, bad = None, [], False
            for pre, addr in terms:
                d = trail(pre)
                if not d:
                    bad = True
                    break
                if dia is None:
                    dia = d
                elif dia != d:
                    bad = True  # 1セルに複数の口径が混ざる
                    break
                sig.append((pre[: -len(d)], addr.replace("$", "")))
            if bad:
                continue

            colinfo.setdefault(cl, {"kei": kei, "hdr": header_of(ws, c.column), "dia": dia})
            for p in sig:
                keys[(kei,) + p].add(c.row)

            gk = (c.row, kei)
            if gk in grp:
                g = grp[gk]
                if g["sig"] != tuple(sig):
                    g["state"] = "NG_SIG"  # 列ごとに参照セルが違う
                g["cols"].append(cl)
            else:
                grp[gk] = {
                    "cols": [cl],
                    "sig": tuple(sig),
                    "tmpl": re.sub(r"'([^']+)" + re.escape(dia) + r"'!", r"'\1#'!", v),
                    "state": "",
                }

    for (_r, kei), g in grp.items():
        if g["state"] == "NG_SIG":
            continue
        shared = set()
        for p in g["sig"]:
            shared |= keys[(kei,) + p]
        # 同じ参照セルを複数の行が分け合う＝行ごとに口径が分かれた区間
        g["state"] = "NG_SHARE" if len(shared) > 1 else "OK"
    return grp, colinfo


def add_siblings(ws, colinfo):
    """リンクのある列と同じ系統の列を、口径が空のまま加える"""
    keis = {v["kei"] for v in colinfo.values() if v["kei"]}
    for c in range(1, ws.max_column + 1):
        cl, k = gl(c), kei_of(ws, c)
        if k in keis and cl not in colinfo:
            colinfo[cl] = {"kei": k, "hdr": header_of(ws, c), "dia": None}


def guess_diameters(wb, colinfo):
    """口径が空の列に、シート名と見出しの数字から候補を当てる"""

    def sheet_exists(kei, dia):
        return any(trail(sh) == dia and norm(kei)[:2] in norm(sh) for sh in wb.sheetnames)

    for _ in range(5):
        changed = False
        for info in colinfo.values():
            if info["dia"]:
                continue
            used = {x["dia"] for x in colinfo.values() if x["kei"] == info["kei"] and x["dia"]}
            cand = {
                n for n in re.findall(r"\d+", info["hdr"])
                if sheet_exists(info["kei"], n) and n not in used
            }
            if len(cand) == 1:
                info["dia"] = cand.pop()
                info["guess"] = True
                changed = True
        if not changed:
            break


def main():
    wbf = openpyxl.load_workbook(BOOK)
    wbv = openpyxl.load_workbook(BOOK, data_only=True)
    ws, wsv = wbf[TARGET], wbv[TARGET]

    grp, colinfo = collect(ws)
    add_siblings(ws, colinfo)
    guess_diameters(wbf, colinfo)

    print("=== 口径対応表 ===")
    for cl in sorted(colinfo, key=lambda x: (len(x), x)):
        i = colinfo[cl]
        mark = "  ★推定" if i.get("guess") else ""
        print(f"   {cl:2}列 口径={str(i['dia']):5} 系統={i['kei']:6} {i['hdr'][:28]}{mark}")

    cols_of_kei = collections.defaultdict(list)
    for cl in sorted(colinfo, key=lambda x: (len(x), x)):
        if colinfo[cl]["dia"]:
            cols_of_kei[colinfo[cl]["kei"]].append(cl)
    print("\n系統 → 数式を入れる列:", dict(cols_of_kei))

    def resolve(tmpl, dia):
        total = 0.0
        for pre, addr in TERM.findall(tmpl.replace("#", dia)):
            if pre not in wbv.sheetnames:
                continue  # シートが無ければ 0（IFERROR 相当）
            v = wbv[pre][addr.replace("$", "")].value
            total += float(v) if isinstance(v, (int, float)) else 0.0
        return total

    same = diff = new_zero = new_nonzero = skipped = 0
    bad = []
    for (r, kei), g in sorted(grp.items()):
        if g["state"] == "NG_SIG":
            skipped += 1
            continue  # 列ごとに参照先が違う行は書き込まない
        targets = cols_of_kei[kei] if g["state"] == "OK" else g["cols"]
        for cl in targets:
            dia = colinfo[cl]["dia"]
            if not dia:
                continue
            got = resolve(g["tmpl"], dia)
            cur = wsv.cell(r, ws[cl + "1"].column).value
            cur = float(cur) if isinstance(cur, (int, float)) else None
            if cl in g["cols"]:
                if cur is not None and abs(got - cur) < 1e-9:
                    same += 1
                else:
                    diff += 1
                    bad.append((r, kei, cl, got, cur))
            elif abs(got) < 1e-9:
                new_zero += 1
            else:
                new_nonzero += 1
                bad.append((r, kei, cl, got, "新規"))

    print(f"\n既存セル : 値が同じ {same} / 変わる {diff}")
    print(f"新規セル : 0が入る {new_zero} / 0以外が入る {new_nonzero}")
    print(f"書き込み対象外（列ごとに参照先が違う）: {skipped} 群")
    for b in bad[:8]:
        print("   ★", b)

    n_ok = sum(1 for g in grp.values() if g["state"] == "OK")
    n_share = sum(1 for g in grp.values() if g["state"] == "NG_SHARE")
    print(f"\n展開可 {n_ok} / 展開不可 {n_share} / 対象外 {skipped}")
    print(f"書き込むセル総数 {same + diff + new_zero + new_nonzero}（うち新規 {new_zero + new_nonzero}）")

    # 既存セルの値が1つでも変わる、または新規セルに0以外が入るなら異常
    return 0 if (diff == 0 and new_nonzero == 0) else 1


if __name__ == "__main__":
    sys.exit(main())
