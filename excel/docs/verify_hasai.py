#!/usr/bin/env python3
"""M_Hasai が書き込む数式を、VBA と同じ手順で再現する。

作成環境に Excel が無いので、マクロの判定をそのまま Python に写して
「どのセルに何が入るか」を出す。VBA を直したらこちらも直すこと。

    python3 excel/docs/verify_hasai.py
    WORK=02_nagata python3 excel/docs/verify_hasai.py
"""
import collections
import os
import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter as gl
from openpyxl.utils import column_index_from_string as ci

HERE = os.path.dirname(os.path.abspath(__file__))
WORK = os.environ.get("WORK", "01_higashishirakawa")
FOLDER = os.path.join(HERE, "..", "works", WORK)
BOOK = os.environ.get("BOOK", "06_dokou_hosou.xlsx")

# ---- M_Hasai の先頭にある設定と同じもの -----------------------------------
TARGET_SHEET = "総括表（土工事）"
COL_MAP = os.environ.get("COL_MAP") or (
    "J=試掘（舗50|K=試掘（舗300|L=試掘（舗75|M=試掘（舗400|N=試掘（舗600|"
    "O=給水2度|P=仮配（舗|Q=給水(舗|"
    "R=管工（舗50|S=管工（舗75|T=管工（舗400|U=管工（舗600")
SECTION_LABEL = "舗装版破砕"
BLOCK_LABEL = "□舗装版破砕"
CUT_LABEL = "舗装切断工"
CUT_BLOCK = "□舗装切断工"
TORI_BLOCK = "□舗装版取壊工"      # 給水2度だけ並びが違う
TORI_KIND = "As"
INPUT_COLOR = "FFFF00"


def norm(v):
    """VBA の Norm と同じ。全角英数を半角に直し、空白を落として大文字に"""
    if v is None:
        return ""
    out = []
    for ch in str(v):
        o = ord(ch)
        if 0xFF01 <= o <= 0xFF5E:
            o -= 0xFEE0
        if o not in (32, 0x3000, 9, 10, 13):
            out.append(chr(o))
    return "".join(out).upper()


def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def merged_value(ws, r, c):
    """VBA の MergeArea.Cells(1,1).Value と同じ"""
    for m in ws.merged_cells.ranges:
        if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
            return ws.cell(m.min_row, m.min_col).value
    return ws.cell(r, c).value


def merged_top(ws, r, c):
    for m in ws.merged_cells.ranges:
        if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
            return m.min_row, m.min_col
    return r, c


def is_input_cell(ws, r, c):
    """黄色く塗ってある入力セルか（VBA の Interior.Color = 65535 と同じ）"""
    f = ws.cell(r, c).fill
    if f is None or f.patternType is None:
        return False
    fg = f.fgColor
    if fg.type == "rgb":
        return isinstance(fg.rgb, str) and fg.rgb.endswith(INPUT_COLOR)
    if fg.type == "indexed":
        return fg.indexed == 13        # 既定パレットの黄色 = FFFF00
    return False


# ---- 転記元シートのブロック ------------------------------------------------
_cache = {}


def sheet_ref(sn):
    """数式に書くシート名。囲む必要のある名前だけ ' で囲む（VBA の SheetRef）"""
    def safe(ch):
        o = ord(ch)
        if ch.isascii() and (ch.isalnum() or ch == "_"):
            return True
        if o in (0x3000, 0x30FB):
            return False
        return (0x3041 <= o <= 0x30FF) or (0x4E00 <= o <= 0x9FFF) or (0xFF66 <= o <= 0xFF9F)

    if not sn or sn[0].isdigit() or not all(safe(c) for c in sn):
        return "'" + sn.replace("'", "''") + "'!"
    return sn + "!"


def merge_wide(ws, r, c):
    for m in ws.merged_cells.ranges:
        if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
            return m.max_col - m.min_col + 1
    return 1


def find_block(wb, sn, anchor=BLOCK_LABEL):
    """転記元のブロックを探し、(r0, r1, 枠の一覧) を返す（VBA の FindBlock）

    枠の一覧は [(種別, 厚さ列, 合計列, 合計の幅), ...]
    """
    key = (anchor, sn)
    if key in _cache:
        return _cache[key]
    _cache[key] = None
    if sn not in wb.sheetnames:
        return None
    ws = wb[sn]
    sect = 0
    for r in range(1, 61):
        for c in range(1, 61):
            if norm(anchor) in norm(ws.cell(r, c).value):
                sect = r
                break
        if sect:
            break
    if not sect:
        return None

    out = scan_kind_header(ws, sect) or scan_gokou_header(ws, sect)
    _cache[key] = out
    return out


def scan_kind_header(ws, sect):
    """並び その1 －「種別・舗装厚」と「合 計」が並ぶ形"""
    hdr, kinds, totals = 0, [], []
    for r in range(sect + 1, sect + 4):
        kinds, totals = [], []
        for c in range(1, 61):
            v = norm(ws.cell(r, c).value)
            if v == norm("種別・舗装厚"):
                kinds.append(c)
            if v == norm("合 計"):
                totals.append(c)
        if kinds and totals:
            hdr = r
            break
    if not hdr:
        return None

    r0, r1 = hdr + 1, hdr
    for r in range(r0, r0 + 41):
        if norm(ws.cell(r, kinds[0]).value) not in ("AS", "CO"):
            break
        r1 = r
    if r1 < r0:
        return None

    pairs = []
    for kc in kinds:
        nxt = [t for t in totals if t > kc]
        if not nxt:
            continue
        sc = min(nxt)
        k = norm(ws.cell(r0, kc).value)
        k = "As" if k == "AS" else ("Co" if k == "CO" else ("As" if not pairs else "Co"))
        pairs.append((k, kc + 1, sc, merge_wide(ws, r0, sc)))
        if len(pairs) == 2:
            break
    if not pairs:
        return None
    return (r0, r1 + 1, pairs)          # 予備を1行


def scan_gokou_header(ws, sect):
    """並び その2 －「車道 5号工」「歩道 5号工」が並ぶ形（給水2度）"""
    hdr, cols = 0, []
    for r in range(sect, sect + 4):
        cols = [c for c in range(1, 61) if norm("号工") in norm(ws.cell(r, c).value)]
        if cols:
            hdr = r
            break
    if not hdr:
        return None

    r0, r1 = hdr + 1, hdr
    for r in range(r0, r0 + 41):
        if ws.cell(r, cols[0]).value is None:
            break
        r1 = r
    if r1 < r0:
        return None

    pairs = []
    for c in cols:
        start, wide = c, merge_wide(ws, hdr, c)
        for m in ws.merged_cells.ranges:
            if m.min_row <= hdr <= m.max_row and m.min_col <= c <= m.max_col:
                start = m.min_col
                break
        pairs.append((TORI_KIND, start, start + 1, max(1, wide - 1)))
    return (r0, r1, pairs) if pairs else None      # 予備は付けない


def rng(col, wide, r0, r1):
    return f"${gl(col)}${r0}:${gl(col + wide - 1)}${r1}"



def hasai_block(wb, sn):
    """舗装版破砕の転記元。給水2度だけ見出しが違うので順に試す"""
    return find_block(wb, sn, BLOCK_LABEL) or find_block(wb, sn, TORI_BLOCK)


def build_sumif(wb, sn, tname, thk_ref, kind):
    b = hasai_block(wb, sn)
    if not b:
        return "", f"{sn} に {BLOCK_LABEL} のブロックがありません"
    r0, r1, pairs = b
    crit = sheet_ref(tname) + thk_ref
    q = sheet_ref(sn)
    terms = []
    for k, thk, tot, w in pairs:
        if k in kind:
            terms.append(f"SUMIF({q}{rng(thk, 1, r0, r1)},{crit},{q}{rng(tot, w, r0, r1)})")
    if not terms:
        return "", f"{sn} に {kind} 側の欄がありません"
    note = ""
    if "Co" in kind and not any(k == "Co" for k, *_ in pairs):
        note = f"{sn} に Co 側の欄が無いため As だけを合計しました"
    return "=" + "+".join(terms), note


def cut_ref(wb, sn, label, kind):
    """舗装切断工の1セル分。行を探して直接参照にする（VBA の CutRef）"""
    b = find_block(wb, sn, CUT_BLOCK)
    if not b:
        return "", f"{sn} に {CUT_BLOCK} のブロックがありません"
    r0, r1, pairs = b
    hit = [(thk, tot) for k, thk, tot, w in pairs if k in kind]
    if not hit:
        return "", f"{sn} に {kind} 側の欄がありません"
    thk, tot = hit[0]
    want = norm_label(label)
    if not want:
        return "", "総括表の厚さ区分が読めません"
    ws = wb[sn]
    for r in range(r0, r1 + 1):
        if norm_label(ws.cell(r, thk).value) == want:
            return "=" + sheet_ref(sn) + f"{gl(tot)}{r}", ""
    return "", f"{sn} に「{label}」の行がありません"


# ---- 総括表側の読み取り ----------------------------------------------------
def section_of(ws, r, first_col):
    """いちばん左の工種名で、どの工種の行かを返す（VBA の SectionOf）"""
    for c in range(1, first_col):
        v = merged_value(ws, r, c)
        if v is None:
            continue
        t = str(v)
        if t.strip() and not t.startswith("="):
            t = norm(t)
            if norm(SECTION_LABEL) in t:
                return SECTION_LABEL
            if norm(CUT_LABEL) in t:
                return CUT_LABEL
            return ""
    return ""


def is_section_row(ws, r, first_col):
    return bool(section_of(ws, r, first_col))


def norm_label(v):
    """厚さ区分の文字をそろえる。t≦15㎝ も t≦15 も同じ（VBA の NormLabel）"""
    return norm(v).replace("\u339d", "").replace("CM", "")


def thick_col(ws, rows, first_col):
    """舗装厚の列。舗装版破砕の行で数値が現れた回数がいちばん多い列（VBA の ThickCol）

    舗装切断工は厚さが区分の文字なので、列を数えるときは見ない。
    """
    hits = collections.Counter()
    for r in rows:
        if section_of(ws, r, first_col) != SECTION_LABEL:
            continue
        for c in range(first_col - 1, 0, -1):
            tr, tc = merged_top(ws, r, c)
            if is_num(ws.cell(tr, tc).value):
                hits[tc] += 1
                break
    return hits.most_common(1)[0][0] if hits else 0


def is_text_cell(ws, r, c):
    """「計」などの文字が入っているか。空欄と数値は False"""
    tr, tc = merged_top(ws, r, c)
    v = ws.cell(tr, tc).value
    if v is None:
        return False
    return isinstance(v, str) and bool(v.strip())


def kind_of_row(ws, r, first_col):
    out = ""
    for c in range(1, first_col):
        t = norm(merged_value(ws, r, c))
        if "AS" in t and "As" not in out:
            out += "As"
        if "CO" in t and "Co" not in out:
            out += "Co"
    return out


def main():
    wb = openpyxl.load_workbook(os.path.join(FOLDER, BOOK))
    if TARGET_SHEET not in wb.sheetnames:
        print("シートがありません:", TARGET_SHEET)
        return 1
    ws = wb[TARGET_SHEET]
    print(f"ブック {WORK}/{BOOK}")

    pairs = [p.split("=", 1) for p in COL_MAP.split("|") if "=" in p]
    first_col = min(ci(c) for c, _ in pairs)

    print("=== 転記元の対応（確認画面に出るもの）===")
    for cl, sn in pairs:
        b = hasai_block(wb, sn)
        if sn not in wb.sheetnames:
            note = "★シートがありません"
        elif not b:
            note = "★破砕のブロックがありません"
        else:
            r0, r1, blk = b
            note = f"破砕 {r0}-{r1}行 " + " ".join(
                f"{k}={gl(t)}/{gl(u)}" + (f"×{w}" if w > 1 else "") for k, t, u, w in blk)
        c = find_block(wb, sn, CUT_BLOCK)
        note += f"  切断 {c[0]}-{c[1]}行" if c else "  切断なし"
        print(f"  {cl}列 → {sn:<12} {note}")

    last = min(ws.max_row, 500)
    rows = [r for r in range(1, last + 1) if is_section_row(ws, r, first_col)]
    print(f"\n{SECTION_LABEL} の行: {len(rows)} 行 "
          f"({rows[0]}〜{rows[-1]})" if rows else "対象行なし")

    tc = thick_col(ws, rows, first_col)
    print(f"舗装厚の列: {gl(tc)} 列")

    written, skipped, notes, sect_of = {}, [], [], {}
    for r in rows:
        sect = section_of(ws, r, first_col)
        istext = is_text_cell(ws, r, tc)
        # 破砕は厚さが数値（「計」の行は触らない）、切断は区分の文字
        if sect == SECTION_LABEL and istext:
            continue
        if sect == CUT_LABEL and not istext:
            continue
        mr, mc = merged_top(ws, r, tc)
        tr = f"${gl(mc)}{mr}"
        kd = kind_of_row(ws, r, first_col)
        if not kd:
            continue
        for cl, sn in pairs:
            if not is_input_cell(ws, r, ci(cl)):
                continue
            if sect == SECTION_LABEL:
                f, note = build_sumif(wb, sn, ws.title, tr, kd)
            else:
                f, note = cut_ref(wb, sn, str(ws.cell(mr, mc).value), kd)
            if not f:
                skipped.append((r, cl, note))
            else:
                written[(r, cl)] = f
                sect_of[(r, cl)] = sect
                if note:
                    notes.append((r, cl, note))

    print(f"\n=== 書き込むセル {len(written)} 個 ===")
    by_row = collections.defaultdict(list)
    for (r, cl) in written:
        by_row[r].append(cl)
    for r in sorted(by_row):
        mr, mc = merged_top(ws, r, tc)
        thk = ws.cell(mr, mc).value
        sect = section_of(ws, r, first_col)
        print(f"  {r:3}行 {sect:<6} 厚さ{str(thk):<9} {kind_of_row(ws, r, first_col):<5} "
              f"→ {','.join(sorted(by_row[r], key=ci))}")

    print("\n=== 舗装切断工（9〜12行）の数式 ===")
    for r in range(9, 13):
        for cl in ("J", "M", "R"):
            if (r, cl) in written:
                print(f"  {cl}{r}: {written[(r, cl)]}")

    print("\n=== 13〜25行 J〜N の数式 ===")
    for r in range(13, 26):
        for cl in ("J", "K", "L", "M", "N"):
            if (r, cl) in written:
                print(f"  {cl}{r}: {written[(r, cl)]}")

    want = ("=SUMIF('試掘（舗50'!$O$11:$O$17,'総括表（土工事）'!$I14,"
            "'試掘（舗50'!$P$11:$P$17)")
    got = written.get((14, "J"), "")
    ok = got == want
    for cell, expect in (
            ("O31", "=SUMIF(給水2度!$K$4:$K$6,'総括表（土工事）'!$I31,給水2度!$L$4:$N$6)"
                    "+SUMIF(給水2度!$O$4:$O$6,'総括表（土工事）'!$I31,給水2度!$P$4:$R$6)"),
            ("J10", "='試掘（舗50'!P5"), ("J12", "='試掘（舗50'!S5"),
                         ("J9", "='試掘（舗50'!P4"), ("M11", "='試掘（舗400'!S4"),
                         ("R9", "='管工（舗50'!T10"), ("S11", "='管工（舗75'!X10")):
        r = int(cell[1:])
        g = written.get((r, cell[0]), "")
        mark = "一致" if g == expect else f"違う（{g}）"
        print(f"{cell} = {expect}  … {mark}")
        ok = ok and g == expect
    print("\nJ14 が指示どおりか:", "一致" if ok else f"違う\n  期待 {want}\n  実際 {got}")

    if skipped:
        print(f"\n見送り {len(skipped)} 個")
        for r, cl, w in skipped[:10]:
            print(f"   {cl}{r}: {w}")
    if notes:
        print(f"\n注意 {len(notes)} 個")
        for r, cl, w in notes[:5]:
            print(f"   {cl}{r}: {w}")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
