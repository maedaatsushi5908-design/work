#!/usr/bin/env python3
"""M_Link が総括表（土工事）に実際に書き込む数式を、VBA と同じ手順で再現する。

Excel の無い環境で「どのセルに何が入るか」を確かめるための道具。
BuildConfig →  WriteAll → FillYellow の順で、VBA と同じ判定をたどる。

    python3 excel/docs/verify_sumif.py

VBA を書き換えたら、こちらも合わせて更新すること。
"""
import collections
import os
import re
import sys
import unicodedata

import openpyxl
from openpyxl.utils import get_column_letter as gl
from openpyxl.utils import column_index_from_string as ci

HERE = os.path.dirname(os.path.abspath(__file__))
BOOK = os.path.join(HERE, "..", "original", "06_dokou_hosou.xlsx")
TARGET = "総括表（土工事）"
TERM = re.compile(r"'([^']+)'!(\$?[A-Z]{1,3}\$?[0-9]+)")


def norm(s):
    if s is None:
        return ""
    t = unicodedata.normalize("NFKC", str(s))
    for ch in " 　\t\r\nφΦfF":
        t = t.replace(ch, "")
    return t.upper()


def trail(s):
    m = re.search(r"(\d+)$", str(s))
    return m.group(1) if m else ""


def header_of(ws, col):
    for r in range(1, 13):
        v = ws.cell(r, col).value
        if v not in (None, "") and norm(v) and not str(v).startswith("="):
            return " ".join(str(v).split())
    return ""


def kei_of(ws, col):
    h = header_of(ws, col)
    if not h:
        return ""
    p = h.find("（")
    if p < 0:
        p = h.find("(")
    if p > 0:
        h = h[:p]
    return h.replace(" ", "").replace("　", "").strip()


# ---- 結合セルを考えた値の取り出し（VBA の MergedValue / KindOf 相当）------
def merged_value(ws, r, c):
    v = ws.cell(r, c).value
    if v not in (None, ""):
        return v
    for m in ws.merged_cells.ranges:
        if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
            return ws.cell(m.min_row, m.min_col).value
    return None


def kind_of(ws, r, c):
    t = norm(merged_value(ws, r, c))
    out = ""
    if "AS" in t:
        out += "As"
    if "CO" in t:
        out += "Co"
    return out


def is_yellow(ws, r, c):
    f = ws.cell(r, c).fill
    if f is None or f.patternType is None:
        return False
    fg = f.fgColor
    if fg.type == "rgb":
        return isinstance(fg.rgb, str) and fg.rgb.endswith("FFFF00")
    if fg.type == "indexed":
        return fg.indexed == 13          # 既定パレットの黄色
    return False


# ---- 舗装版破砕ブロックの検出（VBA の FindBlock 相当）----------------------
_block_cache = {}


def find_block(wb, sn):
    if sn in _block_cache:
        return _block_cache[sn]
    out = None
    if sn in wb.sheetnames:
        ws = wb[sn]
        sect = 0
        for r in range(1, 61):
            for c in range(1, 61):
                if "□舗装版破砕" in norm(ws.cell(r, c).value):
                    sect = r
                    break
            if sect:
                break
        if sect:
            hdr, kinds, totals = 0, [], []
            for r in range(sect + 1, sect + 4):
                kinds, totals = [], []
                for c in range(1, 61):
                    v = norm(ws.cell(r, c).value)
                    if v == "種別・舗装厚":
                        kinds.append(c)
                    if v == "合計":
                        totals.append(c)
                if kinds and totals:
                    hdr = r
                    break
            if hdr:
                pairs = []
                for kc in kinds:
                    nxt = [t for t in totals if t > kc]
                    if nxt:
                        pairs.append((kc + 1, min(nxt)))
                if pairs:
                    a_thk, a_sum = pairs[0]
                    c_thk, c_sum = pairs[1] if len(pairs) > 1 else (0, 0)
                    r0 = hdr + 1
                    r1 = r0 - 1
                    for r in range(r0, r0 + 41):
                        if norm(ws.cell(r, a_thk - 1).value) not in ("AS", "CO"):
                            break
                        r1 = r
                    if r1 >= r0:
                        out = (r0, r1 + 1, a_thk, a_sum, c_thk, c_sum)
    _block_cache[sn] = out
    return out


def rng(col, r0, r1):
    L = gl(col)
    return f"${L}${r0}:${L}${r1}"


def sumif_term(sn, sum_col, thk_col, r0, r1, thk_ref):
    q = f"'{sn}'!"
    return f"SUMIF({q}{rng(thk_col, r0, r1)},{thk_ref},{q}{rng(sum_col, r0, r1)})"


def build_sumif(wb, sn, thk_ref, kind):
    b = find_block(wb, sn)
    if not b:
        return "", f"舗装版破砕工のブロックが見つかりません({sn})"
    r0, r1, a_thk, a_sum, c_thk, c_sum = b
    out = ""
    if "As" in kind and a_sum:
        out = sumif_term(sn, a_sum, a_thk, r0, r1, thk_ref)
    if "Co" in kind and c_sum:
        if out:
            out += "+"
        out += sumif_term(sn, c_sum, c_thk, r0, r1, thk_ref)
    if not out:
        if "Co" in kind and not c_sum:
            return "", f"{sn} に Co 側の欄が無いため埋められません"
        return "", f"種別が As でも Co でもありません({kind})"
    why = ""
    if "Co" in kind and not c_sum:
        why = f"注意: {sn} に Co 側の欄が無いため As だけを合計しています"
    return "=" + out, why


def lead_token(s):
    t = norm(s)
    for ch in ("(", "（"):
        p = t.find(ch)
        if p > 0:
            t = t[:p]
            break
    return t


def sheet_by_kei(wb, kei):
    k = norm(kei)
    if not k:
        return ""
    for sh in wb.sheetnames:
        if norm(sh) == k:
            return sh
    hit, n = "", 0
    for sh in wb.sheetnames:
        tok = lead_token(sh)
        if len(tok) >= 2 and k.startswith(tok) and find_block(wb, sh):
            hit, n = sh, n + 1
    return hit if n == 1 else ""


# ---- BuildConfig 相当 ------------------------------------------------------
def build_config(wb, ws):
    grp, keys, colinfo = {}, collections.defaultdict(set), {}
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not (isinstance(v, str) and v.startswith("=") and "!" in v):
                continue
            if "表紙" in v:
                continue
            terms = TERM.findall(v)
            if not terms or len(terms) != v.count("!"):
                continue
            cl, kei = gl(c.column), kei_of(ws, c.column)
            dia, sig, bad = None, [], False
            for pre, addr in terms:
                d = trail(pre)
                if not d:
                    bad = True
                    break
                if dia is None:
                    dia = d
                elif dia != d:
                    bad = True
                    break
                sig.append((pre[: -len(d)], addr.replace("$", "")))
            if bad:
                continue
            colinfo.setdefault(cl, {"kei": kei, "hdr": header_of(ws, c.column), "dia": dia})
            for p in sig:
                keys[(kei,) + p].add(c.row)
            gk = (c.row, kei)
            if gk in grp:
                g = grp[gk]
                if g["sig"] != tuple(sig):
                    g["state"] = "NG_SIG"
                g["cols"].append(cl)
            else:
                grp[gk] = {"cols": [cl], "sig": tuple(sig), "pre": sig[0][0],
                           "tmpl": v, "state": ""}
    for (_r, kei), g in grp.items():
        if g["state"] == "NG_SIG":
            continue
        shared = set()
        for p in g["sig"]:
            shared |= keys[(kei,) + p]
        g["state"] = "NG_SHARE" if len(shared) > 1 else "OK"

    keis = {v["kei"] for v in colinfo.values() if v["kei"]}
    for c in range(1, ws.max_column + 1):
        cl, k = gl(c), kei_of(ws, c)
        if k in keis and cl not in colinfo:
            colinfo[cl] = {"kei": k, "hdr": header_of(ws, c), "dia": None}

    def sheet_exists(kei, dia):
        return any(trail(sh) == dia and norm(kei)[:2] in norm(sh) for sh in wb.sheetnames)

    for _ in range(5):
        changed = False
        for info in colinfo.values():
            if info["dia"]:
                continue
            used = {x["dia"] for x in colinfo.values() if x["kei"] == info["kei"] and x["dia"]}
            cand = {n for n in re.findall(r"\d+", info["hdr"])
                    if sheet_exists(info["kei"], n) and n not in used}
            if len(cand) == 1:
                info["dia"] = cand.pop()
                info["guess"] = True
                changed = True
        if not changed:
            break
    return grp, colinfo


# ---- WriteAll + FillYellow 相当 -------------------------------------------
def run(ws, wb, thk_col="I", knd_col="E"):
    grp, colinfo = build_config(wb, ws)
    cols_of_kei = collections.defaultdict(list)
    for cl in sorted(colinfo, key=lambda x: (len(x), x)):
        if colinfo[cl]["dia"]:
            cols_of_kei[colinfo[cl]["kei"]].append(cl)

    written = {}          # (row, colletter) -> (formula, 経路, note)
    notes = []

    # --- リンク一覧の行を書く ---
    for (r, kei), g in sorted(grp.items()):
        if g["state"] == "NG_SIG":
            continue
        thk = ws.cell(r, ci(thk_col)).value
        knd = kind_of(ws, r, ci(knd_col))
        first_dia = colinfo[g["cols"][0]]["dia"]
        way = "直接"
        if isinstance(thk, (int, float)) and first_dia and knd:
            if find_block(wb, g["pre"] + first_dia):
                way = "条件式"
        # 条件式は舗装厚で照合するので NG_SHARE でも広げられる
        can_exp = g["state"] == "OK" or (g["state"] == "NG_SHARE" and way == "条件式")
        targets = cols_of_kei[kei] if can_exp else g["cols"]
        for cl in targets:
            dia = colinfo[cl]["dia"]
            if not dia:
                continue
            # 広げる先は黄色い入力セルだけ（VBA の YellowOnly）
            if cl not in g["cols"] and not is_yellow(ws, r, ci(cl)):
                continue
            if way == "条件式":
                f, why = build_sumif(wb, g["pre"] + dia,
                                     f"'{ws.title}'!${thk_col}{r}", knd)
                if not f:
                    notes.append((r, cl, why))
                    continue
            else:
                f = re.sub(r"'([^']+)" + re.escape(first_dia) + r"'!",
                           lambda m: f"'{m.group(1)}{dia}'!", g["tmpl"])
            written[(r, cl)] = (f, "リンク" if cl in g["cols"] else "展開", way)

    # --- 黄色い入力セル ---
    link_rows = [(r, kei, g["pre"]) for (r, kei), g in grp.items()]
    for row in ws.iter_rows():
        for c in row:
            r, cl = c.row, gl(c.column) if not isinstance(c, openpyxl.cell.cell.MergedCell) else None
            if cl is None:
                continue
            if (r, cl) in written:
                continue
            if isinstance(c.value, str) and c.value.startswith("="):
                continue
            if not is_yellow(ws, r, c.column):
                continue
            kei = kei_of(ws, c.column)
            if not kei or cl in (thk_col, knd_col):
                continue          # 厚さ列・種別列・見出しの無い列は転記先ではない
            thk = ws.cell(r, ci(thk_col)).value
            if not isinstance(thk, (int, float)):
                continue
            knd = kind_of(ws, r, ci(knd_col))
            if not knd:
                continue
            sn = ""
            dia = colinfo.get(cl, {}).get("dia")
            if dia:
                best, pre = 10 ** 9, ""
                for lr, lk, lp in link_rows:
                    if norm(lk) == norm(kei) and abs(lr - r) < best:
                        best, pre = abs(lr - r), lp
                if pre:
                    sn = pre + dia
            if not sn:
                sn = sheet_by_kei(wb, kei)
            if not sn:
                notes.append((r, cl, "転記元シートを決められません"))
                continue
            f, why = build_sumif(wb, sn, f"'{ws.title}'!${thk_col}{r}", knd)
            if not f:
                notes.append((r, cl, why))
                continue
            written[(r, cl)] = (f, "黄色", "条件式")
            if why:
                notes.append((r, cl, why))
    return written, notes, colinfo


def main():
    wb = openpyxl.load_workbook(BOOK)
    ws = wb[TARGET]
    written, notes, colinfo = run(ws, wb)

    show_cols = [c for c in ("J", "K", "L", "M", "N") if c in colinfo]
    print("=== 舗装版破砕 13〜25行 ===")
    print(f"{'行':>3} {'厚':>3} {'種別':<4} " + " ".join(f"{c:<8}" for c in show_cols))
    for r in list(range(13, 26)):
        thk = ws.cell(r, 9).value
        knd = kind_of(ws, r, 5)
        cells = []
        for cl in show_cols:
            w = written.get((r, cl))
            if w is None:
                cells.append("-       ")
            else:
                cells.append(f"{w[1]}/{w[2]:<4}"[:8].ljust(8))
        print(f"{r:>3} {str(thk):>3} {knd:<4} " + " ".join(cells))

    print("\n=== 生成された数式（13〜25行 J〜N）===")
    for r in range(13, 26):
        for cl in show_cols:
            w = written.get((r, cl))
            if w:
                print(f"  {cl}{r}: {w[0]}")

    want = "=SUMIF('試掘（舗50'!$O$11:$O$17,'総括表（土工事）'!$I14,'試掘（舗50'!$P$11:$P$17)"
    got = written.get((14, "J"), ("",))[0]
    print("\nJ14 が指示どおりか:", "一致" if got == want else f"違う\n  期待 {want}\n  実際 {got}")

    print(f"\n書き込むセル {len(written)} 個")
    if notes:
        print("埋められない・注意:")
        for r, cl, why in notes[:20]:
            print(f"   {cl}{r}: {why}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
